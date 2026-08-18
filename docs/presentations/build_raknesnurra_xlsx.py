"""Offline räknesnurra — enkel och brandad, för föreningen själv.

Två guldfält: antal medlemmar + snittförsäljning. Stort resultat överst.
Samma formel som roots.nu (35 %). Medlemmar i 5-tal, snitt max 5 000 kr.

    python3 docs/presentations/build_raknesnurra_xlsx.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HERE = Path(__file__).parent
ROOT = Path(__file__).resolve().parents[2]
LOGO = ROOT / "apps/web/public/brand/roots-logo-black.png"
OUT = HERE / "Roots_Raknesnurra.xlsx"
DESKTOP = Path.home() / "Desktop" / "Roots_Raknesnurra.xlsx"

INK = "1D1D1B"
FOREST = "6B794F"
FOREST_SOFT = "EDF1E9"
SAND = "F7F4EE"
SAND_DARK = "7F715B"
SAND_LINE = "D5CABF"
WHITE = "FFFFFF"
INPUT = "FFF6E0"
GOLD = "C4A574"


def fill(hex_rgb: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_rgb)


def font(size=12, bold=False, color=INK, name="Calibri") -> Font:
    return Font(name=name, size=size, bold=bold, color=color)


def thin(color=SAND_LINE) -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def thick_forest() -> Border:
    s = Side(style="medium", color=FOREST)
    return Border(left=s, right=s, top=s, bottom=s)


def paint(ws, r1, r2, c1, c2, hex_rgb):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).fill = fill(hex_rgb)


def build() -> Path:
    wb = Workbook()

    # ── Blad 1: Räkna ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Räkna"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_area = "A1:G32"

    widths = [2.5, 22, 14, 14, 14, 14, 2.5]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    paint(ws, 1, 34, 1, 7, SAND)

    # Toppband
    paint(ws, 1, 4, 1, 7, INK)
    ws.merge_cells("B2:D2")
    ws["B2"] = "ROOTS"
    ws["B2"].font = font(11, bold=True, color=GOLD)
    ws.merge_cells("B3:F3")
    ws["B3"] = "Hur mycket kan er förening få?"
    ws["B3"].font = font(22, bold=True, color=WHITE)
    ws["B3"].alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 36
    ws.row_dimensions[4].height = 8

    if LOGO.exists():
        img = XLImage(str(LOGO))
        img.width = 120
        img.height = 31
        ws.add_image(img, "F2")

    # Steg-för-steg
    ws.merge_cells("B5:F5")
    ws["B5"] = "1. Fyll i de två gula fälten   ·   2. Titta på den gröna siffran   ·   3. Klart"
    ws["B5"].font = font(12, bold=True, color=FOREST)
    ws["B5"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[5].height = 24

    # Input-kort
    paint(ws, 7, 12, 2, 3, WHITE)
    for r in range(7, 13):
        for c in (2, 3):
            ws.cell(r, c).border = thin()

    ws.merge_cells("B7:C7")
    ws["B7"] = "ERA SIFFROR"
    ws["B7"].font = font(10, bold=True, color=FOREST)
    ws["B7"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B7"].fill = fill(FOREST_SOFT)

    ws["B8"] = "Hur många säljer?"
    ws["B8"].font = font(12, bold=True)
    ws["B8"].fill = fill(WHITE)
    ws["C8"] = 50
    ws["C8"].font = font(20, bold=True, color=FOREST)
    ws["C8"].fill = fill(INPUT)
    ws["C8"].border = thick_forest()
    ws["C8"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[8].height = 36

    ws.merge_cells("B9:C9")
    ws["B9"] = "Antal medlemmar som säljer (5, 10, 15 …)"
    ws["B9"].font = font(9, color=SAND_DARK)
    ws["B9"].fill = fill(WHITE)

    ws["B10"] = "Hur mycket säljer var och en?"
    ws["B10"].font = font(12, bold=True)
    ws["B10"].fill = fill(WHITE)
    ws["C10"] = 1500
    ws["C10"].font = font(20, bold=True, color=FOREST)
    ws["C10"].fill = fill(INPUT)
    ws["C10"].border = thick_forest()
    ws["C10"].alignment = Alignment(horizontal="center", vertical="center")
    ws["C10"].number_format = '#,##0 "kr"'
    ws.row_dimensions[10].height = 36

    ws.merge_cells("B11:C11")
    ws["B11"] = "Snitt per person under kampanjen (max 5 000 kr)"
    ws["B11"].font = font(9, color=SAND_DARK)
    ws["B11"].fill = fill(WHITE)

    ws.merge_cells("B12:C12")
    ws["B12"] = '=CONCATENATE("≈ ",ROUND(C10/399,0)," Premiumpaket à 399 kr")'
    ws["B12"].font = font(10, color=FOREST)
    ws["B12"].fill = fill(FOREST_SOFT)
    ws["B12"].alignment = Alignment(horizontal="center")

    # Snabbval
    paint(ws, 7, 12, 5, 6, WHITE)
    for r in range(7, 13):
        for c in (5, 6):
            ws.cell(r, c).border = thin()

    ws.merge_cells("E7:F7")
    ws["E7"] = "SNABBVAL — klicka och kopiera"
    ws["E7"].font = font(10, bold=True, color=FOREST)
    ws["E7"].fill = fill(FOREST_SOFT)
    ws["E7"].alignment = Alignment(horizontal="center")

    ws["E8"] = "Storlek"
    ws["F8"] = "Skriv in →"
    ws["E8"].font = font(9, bold=True, color=SAND_DARK)
    ws["F8"].font = font(9, color=SAND_DARK)
    for r, label, val in ((9, "Litet lag", 20), (10, "Vanligt", 50), (11, "Stort", 100), (12, "Mycket stort", 200)):
        ws.cell(r, 5).value = label
        ws.cell(r, 5).font = font(10)
        ws.cell(r, 5).fill = fill(WHITE)
        ws.cell(r, 6).value = val
        ws.cell(r, 6).font = font(12, bold=True, color=FOREST)
        ws.cell(r, 6).fill = fill(WHITE)
        ws.cell(r, 6).alignment = Alignment(horizontal="center")

    # Hero-resultat
    paint(ws, 14, 18, 2, 6, FOREST)
    ws.merge_cells("B14:F14")
    ws["B14"] = "ER FÖRENING FÅR"
    ws["B14"].font = font(11, bold=True, color="D5CABF")
    ws["B14"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B14"].fill = fill(FOREST)

    ws.merge_cells("B15:F17")
    ws["B15"] = "=ROUND(C8*C10*0.35,0)"
    ws["B15"].font = Font(name="Calibri", size=48, bold=True, color=WHITE)
    ws["B15"].fill = fill(FOREST)
    ws["B15"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B15"].number_format = '#,##0 "kr"'
    ws.row_dimensions[15].height = 28
    ws.row_dimensions[16].height = 28
    ws.row_dimensions[17].height = 28

    ws.merge_cells("B18:F18")
    ws["B18"] = '=CONCATENATE("35 % av ",TEXT(C8*C10,"#,##0")," kr i total försäljning")'
    ws["B18"].font = font(12, color="EDF1E9")
    ws["B18"].fill = fill(FOREST)
    ws["B18"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[18].height = 24

    # Detaljkort
    paint(ws, 20, 22, 2, 6, WHITE)
    details = [
        (2, "Total försäljning", "=C8*C10"),
        (4, "Per säljare till er", "=IF(C8=0,0,ROUND(C8*C10*0.35/C8,0))"),
        (6, "Låst andel", '="35 %"'),
    ]
    # Three columns B-C, D-E conceptually - simpler as three merged pairs
    ws["B20"] = "Total försäljning"
    ws["B20"].font = font(9, color=SAND_DARK)
    ws["B20"].fill = fill(WHITE)
    ws.merge_cells("B21:C21")
    ws["B21"] = "=C8*C10"
    ws["B21"].font = font(18, bold=True)
    ws["B21"].number_format = '#,##0 "kr"'
    ws["B21"].fill = fill(WHITE)
    ws["B21"].alignment = Alignment(horizontal="center")

    ws["D20"] = "Per säljare till er"
    ws["D20"].font = font(9, color=SAND_DARK)
    ws["D20"].fill = fill(WHITE)
    ws.merge_cells("D21:E21")
    ws["D21"] = "=IF(C8=0,0,ROUND(C8*C10*0.35/C8,0))"
    ws["D21"].font = font(18, bold=True)
    ws["D21"].number_format = '#,##0 "kr"'
    ws["D21"].fill = fill(WHITE)
    ws["D21"].alignment = Alignment(horizontal="center")

    ws["F20"] = "Er andel"
    ws["F20"].font = font(9, color=SAND_DARK)
    ws["F20"].fill = fill(WHITE)
    ws["F21"] = "35 %"
    ws["F21"].font = font(18, bold=True, color=FOREST)
    ws["F21"].fill = fill(WHITE)
    ws["F21"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[20].height = 18
    ws.row_dimensions[21].height = 32
    ws.row_dimensions[22].height = 8

    # Exempel-tabell
    ws.merge_cells("B23:F23")
    ws["B23"] = "EXEMPEL NI KAN JÄMFÖRA MED"
    ws["B23"].font = font(10, bold=True, color=FOREST)

    for i, h in enumerate(["", "Medlemmar", "Snitt/person", "Ni får", ""], start=2):
        cell = ws.cell(24, i)
        cell.value = h
        cell.font = font(9, bold=True, color=SAND_DARK)
        cell.fill = fill(WHITE)

    examples = [
        (25, "Litet lag", 20, 1000),
        (26, "Vanlig förening", 50, 1500),
        (27, "Stark kampanj", 100, 2500),
        (28, "Högt snitt", 50, 5000),
    ]
    for row, name, mem, snitt in examples:
        ws.cell(row, 2).value = name
        ws.cell(row, 2).font = font(11)
        ws.cell(row, 2).fill = fill(WHITE)
        ws.cell(row, 3).value = mem
        ws.cell(row, 3).font = font(11)
        ws.cell(row, 3).fill = fill(WHITE)
        ws.cell(row, 3).alignment = Alignment(horizontal="center")
        ws.cell(row, 4).value = snitt
        ws.cell(row, 4).font = font(11)
        ws.cell(row, 4).fill = fill(WHITE)
        ws.cell(row, 4).number_format = '#,##0 "kr"'
        ws.cell(row, 4).alignment = Alignment(horizontal="center")
        ws.cell(row, 5).value = f"=ROUND(C{row}*D{row}*0.35,0)"
        ws.cell(row, 5).font = font(12, bold=True, color=FOREST)
        ws.cell(row, 5).fill = fill(FOREST_SOFT)
        ws.cell(row, 5).number_format = '#,##0 "kr"'
        ws.cell(row, 5).alignment = Alignment(horizontal="center")
        ws.cell(row, 6).fill = fill(WHITE)

    # Tips + footer
    ws.merge_cells("B30:F30")
    ws["B30"] = (
        "Tips: Premiumpaketet kostar 399 kr — då går ca 140 kr till föreningen per paket. "
        "Öppna fliken »Så funkar det« om ni vill förstå mer.  ·  info@roots.nu  ·  roots.nu"
    )
    ws["B30"].font = font(9, color=SAND_DARK)
    ws["B30"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[30].height = 32

    paint(ws, 32, 32, 1, 7, FOREST)
    ws.merge_cells("B32:F32")
    ws["B32"] = "Roots  ·  Räknesnurra för föreningar  ·  35 % låst andel"
    ws["B32"].font = font(9, color=WHITE)
    ws["B32"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B32"].fill = fill(FOREST)
    ws.row_dimensions[32].height = 22

    # Validering
    dv_m = DataValidation(
        type="whole", operator="between", formula1="5", formula2="2000",
        allow_blank=False, showErrorMessage=True, showInputMessage=True,
        promptTitle="Medlemmar", prompt="Välj ett jämnt 5-tal, t.ex. 20, 50 eller 100.",
        errorTitle="Ogiltigt", error="Ange mellan 5 och 2 000 (helst 5-tal).",
    )
    dv_a = DataValidation(
        type="whole", operator="between", formula1="0", formula2="5000",
        allow_blank=False, showErrorMessage=True, showInputMessage=True,
        promptTitle="Snittförsäljning", prompt="Max 5 000 kr per person under kampanjen.",
        errorTitle="För högt", error="Max 5 000 kr per säljare.",
    )
    ws.add_data_validation(dv_m)
    ws.add_data_validation(dv_a)
    dv_m.add(ws["C8"])
    dv_a.add(ws["C10"])

    # Villkorlig varning om ej 5-tal
    ws.conditional_formatting.add(
        "C8",
        FormulaRule(formula=['MOD(C8,5)<>0'], fill=fill("F5D6D0")),
    )

    # ── Blad 2: Så funkar det ────────────────────────────────────────
    help_ws = wb.create_sheet("Så funkar det")
    help_ws.sheet_view.showGridLines = False
    for i, w in enumerate([2.5, 70, 2.5], start=1):
        help_ws.column_dimensions[get_column_letter(i)].width = w
    paint(help_ws, 1, 28, 1, 3, SAND)
    paint(help_ws, 1, 3, 1, 3, INK)

    help_ws.merge_cells("B2:B2")
    help_ws["B2"] = "Så funkar räknesnurran"
    help_ws["B2"].font = font(20, bold=True, color=WHITE)
    help_ws["B2"].fill = fill(INK)
    help_ws.row_dimensions[2].height = 36

    blocks = [
        (5, "Tre steg",
         "1) Skriv hur många som säljer (jämna 5-tal).\n"
         "2) Skriv hur mycket varje person säljer för i snitt.\n"
         "3) Läs av den stora gröna siffran — det är vad föreningen får."),
        (10, "Formeln",
         "Medlemmar × snittförsäljning × 35 % = föreningens förtjänst.\n"
         "Exempel: 50 × 1 500 kr × 35 % = 26 250 kr till er."),
        (14, "Varför 35 %?",
         "Andelen är låst i Roots-erbjudandet — samma siffra i portalen, "
         "avräkningen och den här filen. Ingen förhandling, inga dolda avdrag."),
        (18, "Premiumpaketet",
         "Paketpris 399 kr (schampo + balsam + body wash).\n"
         "Ca 140 kr går till föreningen per sålt paket."),
        (22, "Behöver ni hjälp?",
         "Mejla info@roots.nu eller gå till roots.nu — vi räknar gärna live med er."),
    ]
    for row, title, body in blocks:
        help_ws.cell(row, 2).value = title
        help_ws.cell(row, 2).font = font(13, bold=True, color=FOREST)
        help_ws.cell(row, 2).fill = fill(SAND)
        help_ws.cell(row + 1, 2).value = body
        help_ws.cell(row + 1, 2).font = font(11, color=INK)
        help_ws.cell(row + 1, 2).alignment = Alignment(wrap_text=True, vertical="top")
        help_ws.cell(row + 1, 2).fill = fill(WHITE)
        help_ws.cell(row + 1, 2).border = thin()
        help_ws.row_dimensions[row + 1].height = 55

    wb.save(OUT)
    wb.save(DESKTOP)
    return OUT


if __name__ == "__main__":
    path = build()
    print(f"Wrote {path}")
    print(f"Wrote {DESKTOP}")
