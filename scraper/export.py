"""Excel export with conditional formatting, frozen headers, and hyperlinks."""

from __future__ import annotations

import logging
from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter

from scraper.config import OUTPUT_DIR, SEGMENTS, TOP_N_PRODUCTS
from scraper.models import Product

logger = logging.getLogger(__name__)

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)

COLUMNS = [
    ("Rank", 6),
    ("Produktnamn", 35),
    ("Märke", 20),
    ("Pris (SEK)", 12),
    ("Volym (ml)", 12),
    ("Snittbetyg (1-5)", 16),
    ("Antal recensioner", 18),
    ("Källa", 20),
    ("URL", 40),
    ("Ingrediensförteckning (INCI)", 60),
    ("Naturliga", 10),
    ("Syntetiska", 11),
    ("Totalt", 8),
    ("Passar oss (1-10)", 18),
    ("Motivering", 60),
]


def export_excel(products: list[Product], filename: str = "roots_product_research.xlsx") -> Path:
    output_path = OUTPUT_DIR / filename
    wb = Workbook()
    wb.remove(wb.active)

    for segment, cfg in SEGMENTS.items():
        segment_products = [p for p in products if p.segment == segment]
        segment_products.sort(key=lambda p: p.fit_score or 0, reverse=True)
        top = segment_products[:TOP_N_PRODUCTS]

        ws = wb.create_sheet(title=cfg["display_name"])
        _write_header(ws)
        _write_products(ws, top)
        _apply_formatting(ws, len(top))

    _write_summary_sheet(wb, products)

    wb.save(str(output_path))
    logger.info("Excel saved to %s", output_path)
    return output_path


def _write_header(ws) -> None:
    for col_idx, (name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"


def _write_products(ws, products: list[Product]) -> None:
    for rank, p in enumerate(products, 1):
        row = rank + 1
        ws.cell(row=row, column=1, value=rank)
        ws.cell(row=row, column=2, value=p.product_name)
        ws.cell(row=row, column=3, value=p.brand)
        ws.cell(row=row, column=4, value=p.price_sek)
        ws.cell(row=row, column=5, value=p.volume_ml)
        ws.cell(row=row, column=6, value=p.average_rating)
        ws.cell(row=row, column=7, value=p.num_reviews)
        ws.cell(row=row, column=8, value=p.source)

        url_cell = ws.cell(row=row, column=9, value=p.url)
        if p.url:
            url_cell.hyperlink = p.url
            url_cell.font = Font(color="0563C1", underline="single")

        ws.cell(row=row, column=10, value=p.ingredients_raw[:1000] if p.ingredients_raw else "")
        ws.cell(row=row, column=11, value=p.natural_ingredient_count)
        ws.cell(row=row, column=12, value=p.synthetic_ingredient_count)
        ws.cell(row=row, column=13, value=p.total_ingredient_count)
        ws.cell(row=row, column=14, value=p.fit_score)
        ws.cell(row=row, column=15, value=p.fit_motivation)

        for col in range(1, len(COLUMNS) + 1):
            ws.cell(row=row, column=col).alignment = Alignment(
                vertical="top", wrap_text=(col in (10, 15))
            )


def _apply_formatting(ws, num_products: int) -> None:
    if num_products == 0:
        return
    score_col = get_column_letter(14)
    score_range = f"{score_col}2:{score_col}{num_products + 1}"

    ws.conditional_formatting.add(
        score_range,
        CellIsRule(operator="greaterThanOrEqual", formula=["8"], fill=GREEN_FILL),
    )
    ws.conditional_formatting.add(
        score_range,
        CellIsRule(operator="between", formula=["5", "7.9"], fill=YELLOW_FILL),
    )
    ws.conditional_formatting.add(
        score_range,
        CellIsRule(operator="lessThan", formula=["5"], fill=RED_FILL),
    )


def _write_summary_sheet(wb: Workbook, products: list[Product]) -> None:
    ws = wb.create_sheet(title="Sammanfattning")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15

    bold = Font(bold=True, size=12)
    row = 1

    ws.cell(row=row, column=1, value="Sammanfattning").font = bold
    row += 2

    for segment, cfg in SEGMENTS.items():
        seg_products = [p for p in products if p.segment == segment]
        ws.cell(row=row, column=1, value=cfg["display_name"]).font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value="Antal produkter")
        ws.cell(row=row, column=2, value=len(seg_products))
        row += 1

        prices = [p.price_sek for p in seg_products if p.price_sek]
        if prices:
            ws.cell(row=row, column=1, value="Medelpris (SEK)")
            ws.cell(row=row, column=2, value=round(sum(prices) / len(prices), 0))
            row += 1
            ws.cell(row=row, column=1, value="Lägsta pris (SEK)")
            ws.cell(row=row, column=2, value=round(min(prices), 0))
            row += 1
            ws.cell(row=row, column=1, value="Högsta pris (SEK)")
            ws.cell(row=row, column=2, value=round(max(prices), 0))
            row += 1

        ratings = [p.average_rating for p in seg_products if p.average_rating]
        if ratings:
            ws.cell(row=row, column=1, value="Medelbetyg")
            ws.cell(row=row, column=2, value=round(sum(ratings) / len(ratings), 2))
            row += 1

        scores = [p.fit_score for p in seg_products if p.fit_score]
        if scores:
            ws.cell(row=row, column=1, value="Medel 'Passar oss'")
            ws.cell(row=row, column=2, value=round(sum(scores) / len(scores), 1))
            row += 1

        all_ings: list[str] = []
        for p in seg_products:
            all_ings.extend(p.ingredients_list)
        if all_ings:
            top_ings = Counter(i.lower().strip() for i in all_ings).most_common(10)
            ws.cell(row=row, column=1, value="Vanligaste ingredienser:").font = Font(italic=True)
            row += 1
            for ing_name, count in top_ings:
                ws.cell(row=row, column=1, value=f"  {ing_name}")
                ws.cell(row=row, column=2, value=count)
                row += 1

        source_counts = Counter(p.source for p in seg_products)
        ws.cell(row=row, column=1, value="Produkter per källa:").font = Font(italic=True)
        row += 1
        for src, cnt in source_counts.most_common():
            ws.cell(row=row, column=1, value=f"  {src}")
            ws.cell(row=row, column=2, value=cnt)
            row += 1

        row += 1
